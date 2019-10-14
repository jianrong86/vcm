from django.shortcuts import render, get_object_or_404

# Create your views here.
from vcmapp.models import ReleaseVersion
from datetime import datetime
from django.views.generic.base import View
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.serializers.json import DjangoJSONEncoder
import json
from openpyxl import Workbook
import calendar
from datetime import date, timedelta

data_list = [
    'id',
    'weekly',
    'project_name',
    'customer_name',
    'customer_id',
    'firmware_develop',
    'webui_develop',
    'software_version',
    'software_path',
    'modification',
    'plan_release_date',
    'actual_release_date',
    'release_delay_reason',
    'self_test_result',
    'self_test_fail_reason',
    'val_verify_result',
    'val_verify_fail_reason',
    'create_time',
    'status',
    "compiler",
    "verifier",
    "vpm",
    "build_type",
    "rebuild_reason",
    "release_type",
    "re_release_reason"
]

def export_excel():
    filters = dict()
    wb = Workbook()
    ws = wb.active
    ws.append(data_list)
    software_list = list(ReleaseVersion.objects.filter(**filters).values(*data_list).order_by('id'))
    for item in software_list:
        row_list = list()
        for key in data_list:
            row_list.append(item[key])
        ws.append(row_list)

    wb.save("D:\\sample.xlsx")

def get_month_count():
    filters = dict()
    count = []
    for month in range(1, 13):
        start_date = date.today().replace(month=month, day=1)
        _, days_in_month = calendar.monthrange(start_date.year, start_date.month)
        end_date = start_date + timedelta(days_in_month)
        filters['actual_release_date__range'] = (start_date, end_date)
        month_software = ReleaseVersion.objects.filter(**filters).count()
        count.append(month_software)
    return count

def get_month_rebuild_count():
    filters = dict()
    count = []
    for month in range(1, 13):
        start_date = date.today().replace(month=month, day=1)
        _, days_in_month = calendar.monthrange(start_date.year, start_date.month)
        end_date = start_date + timedelta(days_in_month)
        filters['actual_release_date__range'] = (start_date, end_date)
        filters['build_type'] = '1'
        month_software = ReleaseVersion.objects.filter(**filters).count()
        count.append(month_software)
    return count

def get_month_re_release_count():
    filters = dict()
    count = []
    for month in range(1, 13):
        start_date = date.today().replace(month=month, day=1)
        _, days_in_month = calendar.monthrange(start_date.year, start_date.month)
        end_date = start_date + timedelta(days_in_month)
        filters['actual_release_date__range'] = (start_date, end_date)
        filters['release_type'] = '1'
        month_software = ReleaseVersion.objects.filter(**filters).count()
        count.append(month_software)
    return count

def get_weekly_sw_count():
    filters = dict()
    count = []
    today = date.today()
    weekday = calendar.weekday(today.year, today.month, today.day)
    start_date = today - timedelta(weekday)
    end_date = start_date + timedelta(1)
    for day in range(7):
        filters['actual_release_date__range'] = (start_date, end_date)
        print(start_date, end_date)
        day_count = ReleaseVersion.objects.filter(**filters).count()
        count.append(day_count)
        start_date = end_date
        end_date = end_date + timedelta(1)
    return count

def get_weekly_sw_rebuild_count():
    filters = dict()
    count = []
    today = date.today()
    weekday = calendar.weekday(today.year, today.month, today.day)
    start_date = today - timedelta(weekday)
    end_date = start_date + timedelta(1)
    for day in range(7):
        filters['actual_release_date__range'] = (start_date, end_date)
        filters['build_type'] = '1'
        day_count = ReleaseVersion.objects.filter(**filters).count()
        count.append(day_count)
        start_date = end_date
        end_date = end_date + timedelta(1)
    return count

def get_weekly_sw_re_release_count():
    filters = dict()
    count = []
    today = date.today()
    weekday = calendar.weekday(today.year, today.month, today.day)
    start_date = today - timedelta(weekday)
    end_date = start_date + timedelta(1)
    for day in range(7):
        filters['actual_release_date__range'] = (start_date, end_date)
        filters['release_type'] = '1'
        day_count = ReleaseVersion.objects.filter(**filters).count()
        count.append(day_count)
        start_date = end_date
        end_date = end_date + timedelta(1)
    return count

from openpyxl import load_workbook
def import_excel():
    items = dict()
    wb = load_workbook("E:\\Version Control.xlsx")
    sheet = wb.get_sheet_by_name("版本管理")

    for row in range(2, sheet.max_row + 1):

        items['project_name'] = sheet.cell(row=row, column=2).value
        items['customer_name'] = sheet.cell(row=row, column=3).value
        items['customer_id'] = sheet.cell(row=row, column=4).value
        items['firmware_develop'] = sheet.cell(row=row, column=7).value
        items['webui_develop'] = sheet.cell(row=row, column=8).value
        items['software_version'] = sheet.cell(row=row, column=9).value
        items['software_path'] = sheet.cell(row=row, column=17).value
        items['modification'] = sheet.cell(row=row, column=10).value
        items['plan_release_date'] = "%s 18:00:00" % str(sheet.cell(row=row, column=11).value).replace('.', '-')
        items['actual_release_date'] = "%s 18:00:00" % str(sheet.cell(row=row, column=12).value).replace('.', '-')
        items['release_delay_reason'] = sheet.cell(row=row, column=13).value
        if sheet.cell(row=row, column=14).value == 'PASS':
            items['self_test_result'] = '0'
        elif sheet.cell(row=row, column=14).value == 'FAIL':
            items['self_test_result'] = '1'
        else:
            items['self_test_result'] = '2'
        items['self_test_fail_reason'] = sheet.cell(row=row, column=15).value
        items['status'] = '1'
        items['weekly'] = sheet.cell(row=row, column=1).value

        if sheet.cell(row=row, column=16).value == 'NO':
            items['build_type'] = '0'
        else:
            items['build_type'] = '1'

        if sheet.cell(row=row, column=5).value == 'NO':
            items['release_type'] = '0'
        else:
            items['release_type'] = '1'
        items['re_release_reason'] = sheet.cell(row=row, column=6).value
        print(items)
        ReleaseVersion.objects.create(**items)
        # break

class SoftwareView(View):
    def get(self, request):
        # filters = dict()
        # get_weekly_sw_count()
        # import_excel()
        software_list = ReleaseVersion.objects.all()
        # export_excel()
        # filters = {
        #     'weekly' : 'WK25'
        # }
        # rebuild_filters = {
        #     'weekly': 'WK25',
        #     'build_type' : '1',
        # }
        # re_release_filters = {
        #     'weekly': 'WK25',
        #     'release_type' : '1'
        # }

        month_software_count = get_month_count()
        month_rebuild_count = get_month_rebuild_count()
        month_re_release_count = get_month_re_release_count()
        day_software_count = get_weekly_sw_count()
        day_rebuild_count = get_weekly_sw_rebuild_count()
        day_re_release_count = get_weekly_sw_re_release_count()
        ret = {
            "ver_lists": software_list,
            "month_software_count": month_software_count,
            "month_rebuild_count": month_rebuild_count,
            "month_re_release_count": month_re_release_count,
            "day_software_count":day_software_count,
            "day_rebuild_count":day_rebuild_count,
            "day_re_release_count":day_re_release_count

        }
        return render(request, "index.html", ret)

class SoftwareDelete(View):
    def post(self, request):
        ret = dict(status='success')
        try:
            del_id = request.POST['id']
            print("id:" + del_id)
            ReleaseVersion.objects.get(id=del_id).delete()
        except BaseException as e:
            print("e:"+ str(e))
            ret['status'] = str(e)
        return HttpResponse(json.dumps(ret), content_type='application/json')




class SoftwareList(View):
    def get(self, request):
        fields = ['id', 'weekly','project_name', 'customer_name', 'customer_id', 'software_version',
                  'build_type', 'release_type', 'actual_release_date', 'status']
        filters = dict()
        ret = dict(data=list(ReleaseVersion.objects.filter(**filters).values(*fields).order_by('-id')))
        return HttpResponse(json.dumps(ret, cls=DjangoJSONEncoder), content_type='application/json')

def choice_tuple_to_list(choice_tuple):
    value_list = []
    for tuple_item in choice_tuple:
        value_dict = dict(item=tuple_item[0], value=tuple_item[1])
        value_list.append(value_dict)
    return value_list

class SoftwareCreate(View):
    def get(self,request):
        status_list = choice_tuple_to_list(ReleaseVersion.status_choice)
        test_result_list = choice_tuple_to_list(ReleaseVersion.result_choice)
        build_type_list = choice_tuple_to_list(ReleaseVersion.build_type_choice)
        release_type_list = choice_tuple_to_list(ReleaseVersion.release_type_choice)

        # for status_type in ReleaseVersion.status_choice:
        #     status_dict = dict(item=status_type[0], value=status_type[1])
        #     status_list.append(status_dict)
        # for test_result_type in ReleaseVersion.result_choice:
        #     test_result_dict = dict(item=test_result_type[0], value=test_result_type[1])
        #     test_result_list.append(test_result_dict)
        # for build_type in ReleaseVersion.build_type_choice:
        #     build_type_dict = dict(item=build_type[0], value=build_type[1])
        #     build_type_list.append(build_type_dict)
        # for release_type in ReleaseVersion.release_type_choice:
        #     release_type_dict = dict(item=release_type[0], value=release_type[1])
        #     release_type_list.append(release_type_dict)

        ret = {
            "status_list":status_list,
            "test_result_list":test_result_list,
            "build_type_list":build_type_list,
            "release_type_list":release_type_list
        }
        return render(request, "software_create.html", ret)

    # @csrf_exempt
    def post(self,request):
        ret = dict()
        # if request.is_ajax():
        try:
            plan_date = request.POST.get('plan_release_date')
            print("plan_date")
            print(plan_date)
            if not plan_date:
                plan_date = None

            actual_date = request.POST.get('actual_release_date')
            if not actual_date:
                actual_date = None
            sw_form_post = request.POST.get('software_version')
            try:
                get_ret = list(ReleaseVersion.objects.filter(software_version=sw_form_post))
            except BaseException as e:
                print("get_ret:%s" % e)
            if (get_ret.__len__() == 0):   
                ReleaseVersion.objects.create(
                    project_name=request.POST.get('project_name'),
                    customer_name=request.POST.get('customer_name'),
                    customer_id=request.POST.get('customer_id'),
                    firmware_develop=request.POST.get('firmware_develop'),
                    webui_develop=request.POST.get('webui_develop'),
                    software_version=request.POST.get('software_version'),
                    software_path=request.POST.get('software_path'),
                    modification=request.POST.get('modification'),
                    plan_release_date=plan_date,
                    actual_release_date=actual_date,
                    release_delay_reason=request.POST.get('release_delay_reason'),
                    self_test_result=request.POST.get('self_test_result'),
                    self_test_fail_reason=request.POST.get('self_test_fail_reason'),
                    val_verify_result=request.POST.get('val_verify_result'),
                    val_verify_fail_reason=request.POST.get('val_verify_fail_reason'),
                    create_time=request.POST.get('create_time'),
                    status=request.POST.get('status'),
                    compiler=request.POST.get('compiler'),
                    verifier=request.POST.get('verifier'),
                    vpm=request.POST.get('vpm'),
                    weekly=request.POST.get('weekly'),
                    build_type=request.POST.get('build_type'),
                    rebuild_reason=request.POST.get('rebuild_reason'),
                    release_type=request.POST.get('release_type'),
                    re_release_reason=request.POST.get('re_release_reason')
                )
                ret['status'] = "success"
            else:
                ret['status'] = "fail"
        except BaseException as e:
            ret['status'] = str(e)
        print(ret)
        return HttpResponse(json.dumps(ret), content_type='application/json')


class SoftwareUpdate(View):
    def get(self, request):
        status_list = choice_tuple_to_list(ReleaseVersion.status_choice)
        test_result_list = choice_tuple_to_list(ReleaseVersion.result_choice)
        build_type_list = choice_tuple_to_list(ReleaseVersion.build_type_choice)
        release_type_list = choice_tuple_to_list(ReleaseVersion.release_type_choice)
        # status_list = []
        # test_result_list = []
        # build_type_list = []
        # release_type_list = []
        # for status_type in ReleaseVersion.status_choice:
        #     status_dict = dict(item=status_type[0], value=status_type[1])
        #     status_list.append(status_dict)
        # for test_result_type in ReleaseVersion.result_choice:
        #     test_result_dict = dict(item=test_result_type[0], value=test_result_type[1])
        #     test_result_list.append(test_result_dict)
        # for build_type in ReleaseVersion.build_type_choice:
        #     build_type_dict = dict(item=build_type[0], value=build_type[1])
        #     build_type_list.append(build_type_dict)
        # for release_type in ReleaseVersion.release_type_choice:
        #     release_type_dict = dict(item=release_type[0], value=release_type[1])
        #     release_type_list.append(release_type_dict)
        ver = ReleaseVersion.objects.get(id=request.GET['id'])
        ret = {
            "ver": ver,
            'status_list':status_list,
            'test_result_list':test_result_list,
            "build_type_list": build_type_list,
            "release_type_list": release_type_list
        }
        return render(request, "software_update.html", ret)

    def post(self, request):
        res = dict()
        res['status'] = 'success'
        ver = ReleaseVersion.objects.get(id=request.POST.get('sw_id'))

        if not ver:
            res['status'] = 'fail'

        plan_date = request.POST.get('plan_release_date')
        print("plan_date")
        print(plan_date)
        if not plan_date:
            plan_date = None

        actual_date = request.POST.get('actual_release_date')
        if not actual_date:
            actual_date = None

        try:
            ver.project_name = request.POST.get('project_name')
            ver.customer_name = request.POST.get('customer_name')
            ver.customer_id = request.POST.get('customer_id')
            ver.firmware_develop = request.POST.get('firmware_develop')
            ver.webui_develop = request.POST.get('webui_develop')
            ver.software_version = request.POST.get('software_version')
            ver.software_path = request.POST.get('software_path')
            ver.modification = request.POST.get('modification')
            ver.plan_release_date = plan_date
            ver.actual_release_date = actual_date
            ver.release_delay_reason = request.POST.get('release_delay_reason')
            ver.self_test_result = request.POST.get('self_test_result')
            ver.self_test_fail_reason = request.POST.get('self_test_fail_reason')
            ver.val_verify_result = request.POST.get('val_verify_result')
            ver.val_verify_fail_reason = request.POST.get('val_verify_fail_reason')
            ver.create_time = request.POST.get('create_time')
            ver.status = request.POST.get('status')
            ver.compiler = request.POST.get('compiler')
            ver.verifier = request.POST.get('verifier')
            ver.vpm = request.POST.get('vpm')
            ver.weekly = request.POST.get('weekly')
            ver.build_type = request.POST.get('build_type')
            ver.rebuild_reason = request.POST.get('rebuild_reason')
            ver.release_type = request.POST.get('release_type')
            ver.re_release_reason = request.POST.get('re_release_reason')
            ver.save()
        except BaseException as e:
            res['status'] = str(e)

        return HttpResponse(json.dumps(res), content_type='application/json')


# def new_items(request):
#     # ver_list = ReleaseVersion.objects.all()
#     return render(request, "new_items.html")
#
# def edit_items(request, id):
#     ver = ReleaseVersion.objects.get(id=id)
#     return render(request, "software_update.html", {"ver": ver})
#
# def update_items(request, id):
#     ver = ReleaseVersion.objects.get(id=id)
#
#     if not ver:
#         pass
#
#     try:
#         ver.project_name = request.POST.get('project_name')
#         ver.customer_name = request.POST.get('customer_name')
#         ver.customer_id = request.POST.get('customer_id')
#         ver.firmware_develop = request.POST.get('firmware_develop')
#         ver.webui_develop = request.POST.get('webui_develop')
#         ver.software_version = request.POST.get('software_version')
#         ver.software_path = request.POST.get('software_path')
#         ver.modification = request.POST.get('modification')
#
#         plan_release_time = "%s %s" % (request.POST.get('plan_release_date_0'), request.POST.get('plan_release_date_1'))
#         ver.plan_release_date = datetime.strptime(plan_release_time, "%Y-%m-%d %H:%M%S")
#
#         actual_releaes_time = "%s %s" % (request.POST.get('actual_release_date_0'), request.POST.get('actual_release_date_1'))
#         ver.actual_release_date = datetime.strptime(actual_releaes_time, "%Y-%m-%d %H:%M%S")
#
#         ver.release_delay_reason = request.POST.get('release_delay_reason')
#         ver.self_test_result = request.POST.get('self_test_result')
#         ver.self_test_fail_reason = request.POST.get('self_test_fail_reason')
#         ver.val_verify_result = request.POST.get('val_verify_result')
#         ver.val_verify_fail_reason = request.POST.get('val_verify_fail_reason')
#         ver.create_time = request.POST.get('create_time')
#         ver.status = request.POST.get('status')
#         print("ver.status: %s" % ver.status)
#         ver.save()
#     except BaseException as e:
#         render(request, "404.html")
#
#     return render(request, "software_update.html", {"ver": ver})
#
#
# def submit_items(request):
#     project_name = request.POST.get('project_name')
#
#     ReleaseVersion.objects.create(
#         project_name=request.POST.get('project_name'),
#         customer_name=request.POST.get('customer_name'),
#         customer_id = request.POST.get('customer_id'),
#         firmware_develop = request.POST.get('firmware_develop'),
#         webui_develop = request.POST.get('webui_develop'),
#         software_version = request.POST.get('software_version'),
#         software_path = request.POST.get('software_path'),
#         modification = request.POST.get('modification'),
#         plan_release_date = request.POST.get('plan_release_date'),
#         actual_release_date = request.POST.get('actual_release_date'),
#         release_delay_reason = request.POST.get('release_delay_reason'),
#         self_test_result = request.POST.get('self_test_result'),
#         self_test_fail_reason = request.POST.get('self_test_fail_reason'),
#         val_verify_result = request.POST.get('val_verify_result'),
#         val_verify_fail_reason = request.POST.get('val_verify_fail_reason'),
#         create_time = request.POST.get('create_time'),
#         status = request.POST.get('status'),
#         compiler=request.POST.get('compiler'),
#         verifier = request.POST.get('verifier'),
#         vpm = request.POST.get('vpm')
#     )
#     print("project_name:" + project_name)
#     print("submit_items")
#     ver_list = ReleaseVersion.objects.all()
#     return render(request, 'index.html', {"ver_lists": ver_list})